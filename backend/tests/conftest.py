"""
Fixtures compartidas para pytest.
"""

import warnings
from pathlib import Path

import pytest


@pytest.fixture
def sample_txt_file(tmp_path: Path) -> Path:
    """Crea un archivo .txt de prueba."""
    f = tmp_path / "sample.txt"
    f.write_text("Hello Wizard404. Este es un documento de prueba para búsqueda.")
    return f


@pytest.fixture
def sample_md_file(tmp_path: Path) -> Path:
    """Crea un archivo .md de prueba."""
    f = tmp_path / "readme.md"
    f.write_text("# Título\n\nContenido de markdown con keywords: contrato, oferta.")
    return f


@pytest.fixture
def sample_docx_file(tmp_path: Path) -> Path:
    """Crea un archivo .docx válido con contenido mínimo."""
    from docx import Document

    path = tmp_path / "sample.docx"
    doc = Document()
    doc.add_paragraph("Wizard404 docx test. Contenido extraíble.")
    doc.save(str(path))
    return path


@pytest.fixture
def sample_xlsx_file(tmp_path: Path) -> Path:
    """Crea un archivo .xlsx válido con celdas de texto."""
    from openpyxl import Workbook

    path = tmp_path / "sample.xlsx"
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Wizard404"
        ws["A2"] = "xlsx test"
        ws["B1"] = "extraíble"
        wb.save(str(path))
    return path


@pytest.fixture
def fake_xlsx_file(tmp_path: Path) -> Path:
    """Archivo con extensión .xlsx pero contenido que no es ZIP (reproduce BadZipFile)."""
    f = tmp_path / "fake.xlsx"
    f.write_text("not a zip file")
    return f


@pytest.fixture
def discovery_tree(tmp_path: Path) -> Path:
    """Árbol de directorios con varios tipos de archivo para tests de discovery."""
    (tmp_path / "a.txt").write_text("a")
    (tmp_path / "b.md").write_text("# b")
    (tmp_path / "c.xyz").write_text("c")
    (tmp_path / "fake.xlsx").write_text("not a zip")
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "d.txt").write_text("d")
    (sub / "e.md").write_text("e")
    return tmp_path
