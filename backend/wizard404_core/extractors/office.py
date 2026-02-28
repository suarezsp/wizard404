"""
wizard404_core.extractors.office - Extracción de Office (.docx, .xlsx).

Usa python-docx y openpyxl para extraer texto principal.
"""

import warnings
from pathlib import Path
from datetime import datetime

from wizard404_core.extractors.base import Extractor, register_extractor
from wizard404_core.models import DocumentMetadata


class OfficeExtractor(Extractor):
    """Extractor para .docx y .xlsx."""

    supported_extensions = {".docx", ".xlsx"}

    def extract(self, path: Path) -> DocumentMetadata:
        stat = path.stat()
        try:
            content = self._extract_content(path)
        except Exception:
            return DocumentMetadata(
                path=str(path.resolve()),
                name=path.name,
                mime_type=self._mime_for(path.suffix),
                size_bytes=stat.st_size,
                created_at=datetime.fromtimestamp(stat.st_ctime),
                modified_at=datetime.fromtimestamp(stat.st_mtime),
                content_preview="",
                content_full="",
                is_corrupt=True,
            )
        preview = content[:500].strip() if content else ""
        return DocumentMetadata(
            path=str(path.resolve()),
            name=path.name,
            mime_type=self._mime_for(path.suffix),
            size_bytes=stat.st_size,
            created_at=datetime.fromtimestamp(stat.st_ctime),
            modified_at=datetime.fromtimestamp(stat.st_mtime),
            content_preview=preview,
            content_full=content,
        )

    def _extract_content(self, path: Path) -> str:
        ext = path.suffix.lower()
        if ext == ".docx":
            return self._extract_docx(path)
        if ext == ".xlsx":
            return self._extract_xlsx(path)
        return ""

    def _extract_docx(self, path: Path) -> str:
        from docx import Document

        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _extract_xlsx(self, path: Path) -> str:
        from openpyxl import load_workbook

        # Suprimir UserWarning de openpyxl (unknown extension) durante toda la lectura
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets[:3]:
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" ".join(str(c) for c in row if c))
            return "\n".join(parts)

    def _mime_for(self, ext: str) -> str:
        return {".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}.get(ext, "application/octet-stream")


register_extractor(OfficeExtractor())
